import pandas as pd
import openpyxl

workbook_ar_condicionado = openpyxl.load_workbook('AR- CONDICONADO - PROBLEMAS E QUEIXAS.xlsx', data_only = True)

# Lista para armazenar os dados da aba
data = []
abas_a_pular = ['PROGRAMAÇÃO INICIAL', 'NOVOS INSTALADOS E FUNCIONANDO', 'LISTA DE AR CONDIONADOS', 'QUEIXAS MARÇO E ABRIL 2023', 'ESTOQUE BLOCO K E I', 'Plan2']

for aba in workbook_ar_condicionado.sheetnames:
    if aba in abas_a_pular:
        continue

    planilha = workbook_ar_condicionado[aba]

    # Itera sobre as linhas da aba
    for row in planilha.iter_rows(values_only=True):
        # Adiciona os valores das células da linha à lista de dados
        data.append(row)
    
# Cria um DataFrame a partir dos dados da aba
df_condicionado = pd.DataFrame(data)
df_condicionado = df_condicionado.iloc[:, :8]

# Renomear as colunas
df_condicionado = df_condicionado.rename(columns={
    0: 'Nº CHAMADO',
    1: 'DATA_REGISTRO',
    2: 'SALA',
    3: 'LOCALIZAÇÃO',
    4: 'PROBLEMA',
    5: 'SERVICO',
    6: 'DATA_ABERTURA',
    7: 'DATA_RESOLUCAO',
    8: 'SITUACAO',
})

# Fecha o arquivo após terminar
workbook_ar_condicionado.close()

# Remover linhas vazias
df_condicionado = df_condicionado.dropna()
df_condicionado = df_condicionado[df_condicionado['Nº CHAMADO'] != 'Nº CHAMADO']

df_condicionado = df_condicionado.reset_index(drop=True)

df_condicionado['ABERTURA'] = df_condicionado['Nº CHAMADO'].apply(lambda x: 'AVAC' if 'AVAC' in x else ('WHATSAPP' if 'WHATSAPP' in x else 'DATP'))

df_condicionado['Nº CHAMADO'] = df_condicionado['Nº CHAMADO'].apply(lambda x: str(x).split('#')[1].strip() if '#' in x else x)

df_condicionado.to_excel('Dados Ar Condicionados.xlsx', index=False)