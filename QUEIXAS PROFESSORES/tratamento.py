import pandas as pd
import openpyxl
import os

# Pasta onde estão os arquivos
folder = 'Planilhas/'

# Lista para armazenar os DataFrames de cada arquivo
dfs = []

# Iterar sobre os arquivos na pasta
for filename in os.listdir(folder):
    # Verificar se o arquivo é um arquivo Excel
    if filename.endswith('.xlsx'):
        # Caminho completo do arquivo
        filepath = os.path.join(folder, filename)
        
        # Abrir o arquivo Excel
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        
        # Lista para armazenar os dados de cada aba
        data = []
        
        # Iterar sobre as abas do arquivo
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Iterar sobre as linhas da aba
            for row in sheet.iter_rows(values_only=True):
                # Adicionar os valores das células da linha à lista de dados
                data.append(row)
        
        # Fechar o arquivo
        workbook.close()
        
        # Criar um DataFrame a partir dos dados da aba
        df = pd.DataFrame(data[1:])
        
        # Remover as linhas onde a primeira coluna tem o valor 'DIA'
        df = df[df.iloc[:, 0] != 'DIA']
        
        # Renomear as colunas
        df = df.rename(columns={
            0: 'DIA',
            1: 'HORÁRIO',
            2: 'PROFESSOR (A)',
            3: 'SALA',
            4: 'PROBLEMA RELATADO',
            5: 'AÇÃO REALIZADA',
            6: 'O QUE? (o que vai ser realizado?)',
            7: 'POR QUE? (por que será feito?)',
            8: 'ONDE? (local)',
            9: 'QUEM? (quem é o responsável pela realização?)',
            10: 'QUANDO?',
            11: 'COMO? (como será feito?)'
        })
        
        # Remover linhas vazias
        df = df.dropna()
        
        # Adicionar o DataFrame tratado à lista
        dfs.append(df)

# Concatenar todos os DataFrames em um único DataFrame
df_combined = pd.concat(dfs, ignore_index=True)

# Salvar o DataFrame combinado em um arquivo Excel
df_combined.to_excel('Queixas Professores Tratados.xlsx', index=False)