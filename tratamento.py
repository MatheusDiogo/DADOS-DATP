import pandas as pd
import openpyxl

workbook_queixas = openpyxl.load_workbook('QUEIXAS DOS PROFESSORES- 2023.2.xlsx', data_only = True)
workbook_ar_condicionado = openpyxl.load_workbook('AR- CONDICONADO - PROBLEMAS E QUEIXAS.xlsx', data_only = True)

# Lista para armazenar os dados da aba
data = []
    
for aba in workbook_queixas.sheetnames:
    planilha = workbook_queixas[aba]

    # Itera sobre as linhas da aba
    for row in planilha.iter_rows(values_only=True):
        # Adiciona os valores das células da linha à lista de dados
        data.append(row)

# Fecha o arquivo após terminar
workbook_queixas.close()

# Cria um DataFrame a partir dos dados da aba
df_queixas = pd.DataFrame(data[1:])

# Remover as linhas onde a primeira coluna tem o valor 'DIA'
df_queixas = df_queixas[df_queixas.iloc[:, 0] != 'DIA']

# Renomear as colunas
df_queixas = df_queixas.rename(columns={
    0: 'DIA',
    1: 'HORÁRIO',
    2: 'PROFESSOR (A)',
    3: 'SALA',
    4: 'PROBLEMA RELATADO',
    5: 'AÇÃO REALIZADA',
    6: 'O QUE? (o que vai ser realizado?)',
    7: 'POR QUE? (por que será feito?)',
    8: 'ONDE? (local)',
    9: 'QUEM? (quem é o responsável pela realização?)',
    10: 'QUANDO?',
    11: 'COMO? (como será feito?)'
})

# Remover linhas vazias
df_queixas = df_queixas.dropna()

# Lista para armazenar os dados da aba
data = []
abas_a_pular = ['PROGRAMAÇÃO INICIAL', 'NOVOS INSTALADOS E FUNCIONANDO', 'LISTA DE AR CONDIONADOS', 'QUEIXAS MARÇO E ABRIL 2023', 'ESTOQUE BLOCO K E I']

for aba in workbook_ar_condicionado.sheetnames:
    if aba in abas_a_pular:
        continue

    planilha = workbook_ar_condicionado[aba]

    # Itera sobre as linhas da aba
    for row in planilha.iter_rows(values_only=True):
        # Adiciona os valores das células da linha à lista de dados
        data.append(row)
    
# Cria um DataFrame a partir dos dados da aba
df_condicionado = pd.DataFrame(data)
    
# Fecha o arquivo após terminar
workbook_queixas.close()

df_queixas.to_excel('Queixas Professores Tratados.xlsx', index=False)
df_condicionado.to_excel('Dados Ar Condicionados.xlsx', index=False)